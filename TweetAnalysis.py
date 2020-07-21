import re
from collections import defaultdict, Counter
import Algorithmia
import tweepy
import openpyxl



client = Algorithmia.client("api_key")												## API key ##



## function for retrieving 1000 tweets ##
def pull_tweets():
    input = {
        "query": "coronavirus",                                                        ## Keyword for searching ##
        "numTweets": "1000",                                                             ## Number of tweets ##
        "auth": {
            "app_key": 'consumer_key',                                      ## Twitter keys ##
            "app_secret": 'consumer_secret_key',
            "oauth_token": 'access_token',
            "oauth_token_secret": 'access_token_secret'
        }
    }
    
    twitter_algo = client.algo("twitter/RetrieveTweetsWithKeyword/0.1.3")
    result = twitter_algo.pipe(input).result
    tweet_list = [tweets['text'] for tweets in result]

    workbook = openpyxl.Workbook()                                                  ## Excel sheet for entering tweets ##
    sheet = workbook.active
    sheet.title = "Tweets"
    cell1 = sheet.cell(row=1, column=1)
    cell1.value = "Timestamp"
    cell1 = sheet.cell(row=1, column=2)
    cell1.value = "Tweet"
    for i in range(len(result)):
        c1 = sheet.cell(row=i+2, column=1)
        c1.value = result[i]['created_at']
        c2 = sheet.cell(row=i+2, column=2)
        c2.value = result[i]['text']
    workbook.save(r"your_file_path")    		## Add the project older path here Eg. C:\Users\Ankita\Desktop\Coronavirus\InputTweets.xlsx  (InputTweets.xlsx is the file name) ##
    return tweet_list


## function for preprocessing tweets ##
def process_text():
    """Remove emoticons, numbers etc. and returns list of cleaned tweets."""
    data = pull_tweets()
    regex_remove = "(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)|^RT|http.+?"
    stripped_text = [
        re.sub(regex_remove, '',
               tweets).strip() for tweets in data
    ]
    return '. '.join(stripped_text)


## function for performing Named Entity Recognition (NER) ##
def get_ner():
    """Get named entities from the NER algorithm using clean tweet data."""
    data = process_text()
    ner_algo = client.algo(
        'StanfordNLP/NamedEntityRecognition/0.1.1').set_options(timeout=600)
    ner_result = ner_algo.pipe(data).result
    return ner_result


## function for getting counts of each location and organization ##
def group_data():
    data = get_ner()
    default_dict = defaultdict(list)
    for items in data:
        for k, v in items:
            if 'LOCATION' in v or 'ORGANIZATION' in v or 'NAME' in v:
                default_dict[v].append(k)
    ner_list = [{keys: Counter(values)}
            for (keys, values) in default_dict.items()]
    return ner_list


## MAIN FUNCTION ##
if(__name__ == '__main__'):
    result = group_data()
    data = result[0]
    for i in data.keys():
        data[i] = dict(data[i])
    if('LOCATION' in data.keys()):
        locations = data['LOCATION']
        f = open('Output.txt','w')                                                         ## Text file for saving output ##
        if(len(locations)>=10):
            print("Top 10 places with highest number of tweets:-\n")
            f.write("Top 10 places with highest number of tweets:-\n")
            for _ in range(10):
                place = max(locations.items(), key = lambda x: x[1])
                print("{0}: {1}".format(place[0], place[1]))
                f.write(place[0]+": "+str(place[1])+"\n")
                del locations[place[0]]
        else:
            print("Places with highest number of tweets:-\n")
            f.write("Places with highest number of tweets:-\n")
            for place in locations:
                print("{0}: {1}".format(place, locations[place]))
                f.write(place+": "+str(locations[place])+"\n")
        f.close()